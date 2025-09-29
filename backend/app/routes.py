from flask import Blueprint, request, jsonify
from . import db
from .services import get_face_encoding, match_face
from .utils import role_required
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity
import bcrypt
from bson import ObjectId
import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from flask import send_file

api_bp = Blueprint('api', __name__)

# --- AUTH ROUTES ---
@api_bp.route('/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    identifier = data.get('identifier')
    password = data.get('password')
    user = db.users.find_one({"$or": [{"email": identifier}, {"roll_no": identifier}]})
    if user and bcrypt.checkpw(password.encode('utf-8'), user['password']):
        user_id = str(user['_id'])
        role = user['role']
        additional_claims = {"role": role}
        access_token = create_access_token(identity=user_id, additional_claims=additional_claims)
        return jsonify(access_token=access_token, role=role, name=user.get('name'))
    return jsonify({"msg": "Bad username or password"}), 401


# --- ADMIN ROUTES ---
@api_bp.route('/admin/register-student', methods=['POST'])
@role_required('admin')
def register_student():
    name = request.form.get('name')
    roll_no = request.form.get('roll_no')
    course_id = request.form.get('course_id')
    password = request.form.get('password')
    if 'face_image' not in request.files: return jsonify(msg="No face image provided"), 400
    face_image = request.files['face_image']
    if db.users.find_one({"roll_no": roll_no}): return jsonify(msg="Student with this roll number already exists"), 409
    face_encoding = get_face_encoding(face_image)
    if face_encoding is None: return jsonify(msg="Could not detect a single face in the image."), 400
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    user_doc = {"name": name, "roll_no": roll_no, "password": hashed_password, "role": "student", "course_id": course_id, "face_encoding": face_encoding}
    db.users.insert_one(user_doc)
    return jsonify(msg="Student registered successfully"), 201

@api_bp.route('/admin/analytics', methods=['GET'])
@role_required('admin')
def get_admin_analytics():
    course_ids_str = request.args.get('courses')
    selected_course_ids = course_ids_str.split(',') if course_ids_str else []
    
    # --- SETUP FILTERS ---
    student_filter = {"role": "student"}
    course_filter = {}
    
    if selected_course_ids:
        student_filter["course_id"] = {"$in": selected_course_ids}
        # Handle querying courses by _id, which is likely an ObjectId
        try:
            object_ids = [ObjectId(cid) for cid in selected_course_ids]
            course_filter["_id"] = {"$in": object_ids}
        except Exception:
            # Fallback for non-ObjectId strings (e.g., 'CS101')
            course_filter["_id"] = {"$in": selected_course_ids}
            
    # --- OVERALL ANALYTICS (Corrected) ---
    
    # 1. Get the list of *currently existing* student IDs. This is the source of truth.
    current_students_cursor = db.users.find(student_filter, {"_id": 1})
    current_student_ids = [s['_id'] for s in current_students_cursor]
    total_students = len(current_student_ids)

    if total_students == 0:
        return jsonify({"totalStudents": 0, "overallAttendancePercentage": 0, "courseAnalytics": []})

    # 2. Build attendance filter based ONLY on these existing students.
    attendance_filter = {"student_id": {"$in": current_student_ids}}
    if selected_course_ids:
        attendance_filter["course_id"] = {"$in": selected_course_ids}
        
    # 3. Perform calculations using the consistent student list.
    total_attendance_records = db.attendance.count_documents(attendance_filter)
    distinct_dates = db.attendance.distinct("date", attendance_filter)
    total_possible_attendance = total_students * len(distinct_dates)
    overall_percentage = (total_attendance_records / total_possible_attendance) * 100 if total_possible_attendance > 0 else 0

    # --- COURSE-BY-COURSE ANALYTICS (Corrected) ---
    courses_to_analyze = list(db.courses.find(course_filter if selected_course_ids else {}))
    
    course_analytics = []
    for course in courses_to_analyze:
        course_id_str = str(course['_id'])
        
        # Get current students for THIS specific course.
        students_in_course_cursor = db.users.find({"role": "student", "course_id": course_id_str}, {"_id": 1})
        student_ids_in_course = [s['_id'] for s in students_in_course_cursor]
        students_in_course_count = len(student_ids_in_course)

        if students_in_course_count == 0:
            course_analytics.append({"name": course['name'], "attendance": 0})
            continue

        # Filter attendance by THIS course's current students.
        course_attendance_filter = {
            "course_id": course_id_str,
            "student_id": {"$in": student_ids_in_course}
        }
        
        attendance_in_course = db.attendance.count_documents(course_attendance_filter)
        distinct_dates_in_course = db.attendance.distinct("date", course_attendance_filter)
        
        possible_in_course = students_in_course_count * len(distinct_dates_in_course)
        course_percentage = (attendance_in_course / possible_in_course) * 100 if possible_in_course > 0 else 0
        
        course_analytics.append({"name": course['name'], "attendance": round(course_percentage, 2)})

    return jsonify({
        "totalStudents": total_students, 
        "overallAttendancePercentage": round(overall_percentage, 2), 
        "courseAnalytics": course_analytics
    })


# --- TEACHER ROUTES ---
@api_bp.route('/teacher/mark-attendance', methods=['POST'])
@role_required('teacher')
def mark_attendance():
    course_id = request.form.get('course_id')
    if 'live_image' not in request.files: return jsonify(msg="No image captured"), 400
    live_image = request.files['live_image']
    students = list(db.users.find({"role": "student", "course_id": course_id, "face_encoding": {"$exists": True}}))
    if not students: return jsonify(msg="No students with face data for this course"), 404
    known_encodings = [s['face_encoding'] for s in students]
    match_index = match_face(known_encodings, live_image.stream)
    if match_index is not None:
        matched_student = students[match_index]
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        db.attendance.update_one(
            {"student_id": matched_student['_id'], "course_id": course_id, "date": today}, 
            {"$set": {"status": "Present"}}, 
            upsert=True
        )
        return jsonify(msg=f"Attendance marked for {matched_student['name']}"), 200
    else:
        return jsonify(msg="No match found."), 404


@api_bp.route('/teacher/analytics/<course_id>', methods=['GET'])
@role_required('teacher')
def get_teacher_course_analytics(course_id):
    month_str = request.args.get('month', datetime.datetime.now().strftime("%Y-%m"))

    # 1. Get the list of *currently existing* student IDs for this course.
    current_students_cursor = db.users.find({"role": "student", "course_id": course_id}, {"_id": 1})
    current_student_ids = [s['_id'] for s in current_students_cursor]
    students_in_course_count = len(current_student_ids)

    if students_in_course_count == 0:
        return jsonify({
            "studentsInCourse": 0,
            "overallAttendancePercentage": 0,
            "dailyStats": [],
            "totalClasses": 0
        })

    # 2. Build filters based ONLY on these existing students.
    attendance_filter = {
        "course_id": course_id,
        "student_id": {"$in": current_student_ids},
        "date": {"$regex": f"^{month_str}"}
    }
    
    # 3. Recalculate monthly percentage correctly.
    total_attendance_records = db.attendance.count_documents(attendance_filter)
    distinct_dates_for_month = db.attendance.distinct("date", attendance_filter)
    
    total_possible_attendance = students_in_course_count * len(distinct_dates_for_month)
    overall_percentage = (total_attendance_records / total_possible_attendance) * 100 if total_possible_attendance > 0 else 0

    # 4. Correct the daily stats for the last 7 sessions.
    all_class_dates = db.attendance.distinct("date", {"course_id": course_id, "student_id": {"$in": current_student_ids}})
    recent_dates = sorted(list(all_class_dates), reverse=True)[:7]

    daily_stats = []
    for date_str in recent_dates:
        # Count "Present" records for *current* students on that specific day.
        present_count = db.attendance.count_documents({
            "course_id": course_id, 
            "date": date_str, 
            "status": "Present",
            "student_id": {"$in": current_student_ids}
        })
        # Absent count is now correct and can no longer be negative.
        absent_count = students_in_course_count - present_count
        daily_stats.append({"date": date_str, "Present": present_count, "Absent": absent_count})

    return jsonify({
        "studentsInCourse": students_in_course_count,
        "overallAttendancePercentage": round(overall_percentage, 2),
        "dailyStats": daily_stats,
        "totalClasses": len(distinct_dates_for_month)
    })

# --- STUDENT ROUTE ---
@api_bp.route('/student/attendance', methods=['GET'])
@jwt_required()
def get_student_attendance():
    current_user_id = get_jwt_identity()
    try:
        student_obj_id = ObjectId(current_user_id)
    except Exception:
        return jsonify(msg="Invalid student ID in token"), 400
    
    student = db.users.find_one({"_id": student_obj_id})
    if not student: return jsonify(msg="Student not found"), 404
    
    view_mode = request.args.get('view', 'daily')
    student_course_id = student.get("course_id")
    
    all_session_dates = sorted(db.attendance.distinct("date", {"course_id": student_course_id}))
    student_present_dates = {rec['date'] for rec in db.attendance.find({"student_id": student_obj_id, "status": "Present"})}
    
    if view_mode == 'weekly':
        weekly_stats = defaultdict(lambda: {"sessions": 0, "present": 0})
        for date_str in all_session_dates:
            try:
                dt_obj = datetime.datetime.strptime(date_str.strip(), "%Y-%m-%d")
                year, week_num, _ = dt_obj.isocalendar()
                week_key = f"{year}-W{week_num:02d}"
                weekly_stats[week_key]["sessions"] += 1
                if date_str in student_present_dates: weekly_stats[week_key]["present"] += 1
            except (ValueError, TypeError):
                continue
        timeline = []
        for week, stats in sorted(weekly_stats.items()):
            percentage = round((stats["present"] / stats["sessions"]) * 100, 2) if stats["sessions"] > 0 else 0
            timeline.append({"week": week, "percentage": percentage})
        return jsonify({"viewData": timeline, "studentName": student.get('name')})
    
    elif view_mode == 'monthly':
        month_str = request.args.get('month')
        if not month_str: return jsonify(msg="Month parameter is required"), 400
        present_count, absent_count = 0, 0
        for date_str in all_session_dates:
            if date_str and date_str.startswith(month_str):
                if date_str in student_present_dates:
                    present_count += 1
                else:
                    absent_count += 1
        pie_data = [{"name": "Present", "value": present_count}, {"name": "Absent", "value": absent_count}]
        return jsonify({"viewData": pie_data, "studentName": student.get('name')})
        
    elif view_mode == 'daily':
        daily_log = []
        for date_str in sorted(all_session_dates, reverse=True):
            status = "Present" if date_str in student_present_dates else "Absent"
            daily_log.append({"date": date_str, "status": status})
        return jsonify({"viewData": daily_log, "studentName": student.get('name')})
        
    return jsonify(msg="Invalid view mode"), 400


# --- SHARED & OTHER ADMIN ROUTES ---
@api_bp.route('/courses', methods=['GET'])
@jwt_required()
def get_courses():
    courses = list(db.courses.find({}))
    for course in courses:
        course['_id'] = str(course['_id'])
    return jsonify(courses)

@api_bp.route('/students', methods=['GET'])
@role_required('admin')
def get_all_students():
    students = list(db.users.find({"role": "student"}, {"name": 1, "roll_no": 1, "course_id": 1}))
    for student in students:
        student['_id'] = str(student['_id'])
    return jsonify(students)

@api_bp.route('/admin/student-analytics/<student_id>', methods=['GET'])
@role_required('admin')
def get_student_analytics(student_id):
    try:
        student_obj_id = ObjectId(student_id)
    except Exception:
        return jsonify(msg="Invalid student ID format"), 400
    
    view_mode = request.args.get('view', 'weekly')
    student = db.users.find_one({"_id": student_obj_id})
    if not student: return jsonify(msg="Student not found"), 404
    
    student_course_id = student.get("course_id")
    all_session_dates = sorted(db.attendance.distinct("date", {"course_id": student_course_id}))
    student_present_dates = {rec['date'] for rec in db.attendance.find({"student_id": student_obj_id, "status": "Present"})}
    
    if view_mode == 'weekly':
        weekly_stats = defaultdict(lambda: {"sessions": 0, "present": 0})
        for date_str in all_session_dates:
            try:
                dt_obj = datetime.datetime.strptime(date_str.strip(), "%Y-%m-%d")
                year, week_num, _ = dt_obj.isocalendar()
                week_key = f"{year}-W{week_num:02d}"
                weekly_stats[week_key]["sessions"] += 1
                if date_str in student_present_dates: weekly_stats[week_key]["present"] += 1
            except (ValueError, TypeError):
                continue
        timeline = []
        for week, stats in sorted(weekly_stats.items()):
            percentage = round((stats["present"] / stats["sessions"]) * 100, 2) if stats["sessions"] > 0 else 0
            timeline.append({"week": week, "percentage": percentage})
        return jsonify(timeline)
        
    elif view_mode == 'monthly':
        month_str = request.args.get('month')
        if not month_str: return jsonify(msg="Month parameter is required for monthly view"), 400
        present_count, absent_count = 0, 0
        for date_str in all_session_dates:
            if date_str and date_str.startswith(month_str):
                if date_str in student_present_dates:
                    present_count += 1
                else:
                    absent_count += 1
        return jsonify([{"name": "Present", "value": present_count}, {"name": "Absent", "value": absent_count}])
        
    elif view_mode == 'daily':
        daily_log = []
        for date_str in sorted(all_session_dates, reverse=True):
            status = "Present" if date_str in student_present_dates else "Absent"
            daily_log.append({"date": date_str, "status": status})
        return jsonify(daily_log)
        
    return jsonify(msg="Invalid view mode specified"), 400

@api_bp.route('/profile/me', methods=['GET'])
@jwt_required()
def get_my_profile():
    current_user_id = get_jwt_identity()
    user = db.users.find_one({"_id": ObjectId(current_user_id)}, {"password": 0})
    if not user:
        return jsonify(msg="User not found"), 404

    user['_id'] = str(user['_id'])
    
    if user.get('role') == 'student' and 'course_id' in user:
        try:
            # course_id in user collection can be string or ObjectId, handle both
            course_obj_id = ObjectId(user['course_id'])
            course = db.courses.find_one({"_id": course_obj_id})
        except Exception:
            course = db.courses.find_one({"_id": user['course_id']})
            
        if course:
            user['course_name'] = course.get('name')

    return jsonify(user), 200

@api_bp.route('/admin/student/<student_id>', methods=['PUT'])
@role_required('admin')
def update_student(student_id):
    try:
        student_obj_id = ObjectId(student_id)
        data = request.get_json()
        update_data = {}
        if 'name' in data: update_data['name'] = data['name']
        if 'roll_no' in data: update_data['roll_no'] = data['roll_no']
        if 'course_id' in data: update_data['course_id'] = data['course_id']
        if 'password' in data and data['password']:
            hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
            update_data['password'] = hashed_password
        
        if not update_data:
            return jsonify(msg="No update data provided"), 400
        
        result = db.users.update_one({"_id": student_obj_id, "role": "student"}, {"$set": update_data})
        if result.matched_count == 0:
            return jsonify(msg="Student not found"), 404
            
        return jsonify(msg="Student updated successfully"), 200
    except Exception as e:
        return jsonify(msg=f"An error occurred: {str(e)}"), 500

@api_bp.route('/admin/student/<student_id>', methods=['DELETE'])
@role_required('admin')
def delete_student(student_id):
    try:
        student_obj_id = ObjectId(student_id)
        result = db.users.delete_one({"_id": student_obj_id, "role": "student"})
        if result.deleted_count == 0:
            return jsonify(msg="Student not found"), 404
        # Also delete associated attendance records
        db.attendance.delete_many({"student_id": student_obj_id})
        return jsonify(msg="Student and their attendance records deleted successfully"), 200
    except Exception as e:
        return jsonify(msg=f"An error occurred: {str(e)}"), 500
    
@api_bp.route('/teacher/report/<course_id>', methods=['GET'])
@role_required('teacher')
def get_course_report(course_id):
    month_str = request.args.get('month')
    if not month_str:
        return jsonify(msg="Month parameter is required"), 400

    students = list(db.users.find({"role": "student", "course_id": course_id}, {"name": 1, "roll_no": 1}))
    
    date_filter = {"course_id": course_id, "date": {"$regex": f"^{month_str}"}}
    distinct_dates = sorted(db.attendance.distinct("date", date_filter))

    wb = Workbook()
    ws = wb.active
    
    try:
        course_name = db.courses.find_one({"_id": ObjectId(course_id)})['name']
    except Exception:
        course_name = db.courses.find_one({"_id": course_id})['name']

    ws.title = f"{course_name} Attendance"

    headers = ["Roll No", "Student Name"] + distinct_dates
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for student in students:
        student_id = student['_id']
        row_data = [student.get('roll_no', 'N/A'), student.get('name', 'N/A')]
        
        present_dates = set(rec['date'] for rec in db.attendance.find({
            "student_id": student_id,
            "date": {"$in": distinct_dates},
            "status": "Present"
        }))

        for date in distinct_dates:
            status = "P" if date in present_dates else "A"
            row_data.append(status)
        ws.append(row_data)

    mem_file = BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)

    return send_file(
        mem_file,
        as_attachment=True,
        download_name=f'attendance_report_{course_name}_{month_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@api_bp.route('/admin/full-report', methods=['GET'])
@role_required('admin')
def get_full_school_report():
    month_str = request.args.get('month')
    course_ids_str = request.args.get('courses')

    if not month_str or not course_ids_str:
        return jsonify(msg="Month and course IDs are required"), 400

    course_ids = course_ids_str.split(',')
    
    try:
        course_obj_ids = [ObjectId(cid) for cid in course_ids]
        courses = list(db.courses.find({"_id": {"$in": course_obj_ids}}))
    except Exception:
        courses = list(db.courses.find({"_id": {"$in": course_ids}}))

    students = list(db.users.find({"role": "student", "course_id": {"$in": course_ids}}, {"name": 1, "roll_no": 1, "course_id": 1}))
    
    wb = Workbook()
    wb.remove(wb.active)

    for course in courses:
        course_id_str = str(course['_id'])
        course_name = course['name']
        ws = wb.create_sheet(title=course_name[:31])

        date_filter = {"course_id": course_id_str, "date": {"$regex": f"^{month_str}"}}
        distinct_dates = sorted(db.attendance.distinct("date", date_filter))

        headers = ["Roll No", "Student Name"] + distinct_dates
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        students_in_course = [s for s in students if s.get('course_id') == course_id_str]

        for student in students_in_course:
            student_id = student['_id']
            row_data = [student.get('roll_no', 'N/A'), student.get('name', 'N/A')]
            
            present_dates = set(rec['date'] for rec in db.attendance.find({
                "student_id": student_id,
                "date": {"$in": distinct_dates},
                "status": "Present"
            }))
            
            for date in distinct_dates:
                status = "P" if date in present_dates else "A"
                row_data.append(status)
            
            ws.append(row_data)

    mem_file = BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)

    return send_file(
        mem_file,
        as_attachment=True,
        download_name=f'school_report_{month_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@api_bp.route('/admin/student/<student_id>/update-face', methods=['POST'])
@role_required('admin')
def update_student_face(student_id):
    try:
        student_obj_id = ObjectId(student_id)
        if 'face_image' not in request.files:
            return jsonify(msg="No face image provided"), 400
        
        face_image = request.files['face_image']
        face_encoding = get_face_encoding(face_image)

        if face_encoding is None:
            return jsonify(msg="Could not detect a single face in the image."), 400

        result = db.users.update_one(
            {"_id": student_obj_id, "role": "student"},
            {"$set": {"face_encoding": face_encoding}}
        )

        if result.matched_count == 0:
            return jsonify(msg="Student not found"), 404
            
        return jsonify(msg="Student face image updated successfully"), 200
    except Exception as e:
        return jsonify(msg=f"An error occurred: {str(e)}"), 500